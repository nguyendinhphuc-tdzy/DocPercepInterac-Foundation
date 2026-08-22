import sys
from pathlib import Path
sys.path.insert(0, '.')
sys.path.insert(0, './foundation')

import openpyxl
from tests.evaluation.rollforward_profiler import PATH_DATA_FARPT, PATH_DATA_APP1

print(f"Loading FA&RPT: {PATH_DATA_FARPT}")
wb_farpt = openpyxl.load_workbook(str(PATH_DATA_FARPT), data_only=True)
print(f"FA&RPT Sheets: {wb_farpt.sheetnames}")

print(f"\nLoading Appendix I: {PATH_DATA_APP1}")
wb_app1 = openpyxl.load_workbook(str(PATH_DATA_APP1), data_only=True)
print(f"Appendix I Sheets: {wb_app1.sheetnames}")

print("\n--- FA&RPT 'FS' (Rows 1 to 20) ---")
ws_fs = wb_farpt["FS"]
for r in range(1, 20):
    vals = [ws_fs.cell(r, c).value for c in range(1, 8)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")

print("\n--- FA&RPT 'Financial Analysis' (Rows 1 to 36) ---")
ws_fa = wb_farpt["Financial Analysis"]
for r in range(1, 36):
    vals = [ws_fa.cell(r, c).value for c in range(1, 6)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")

if "TP" in wb_farpt.sheetnames or "Benchmarking" in wb_farpt.sheetnames or "RPTs" in wb_farpt.sheetnames:
    for sname in ["RPTs", "TP", "Benchmarking", "Search"]:
        if sname in wb_farpt.sheetnames:
            print(f"\n--- FA&RPT '{sname}' (Rows 1 to 25) ---")
            ws = wb_farpt[sname]
            for r in range(1, 25):
                vals = [ws.cell(r, c).value for c in range(1, 8)]
                if any(vals):
                    print(f"Row {r:2d}: {vals}")
