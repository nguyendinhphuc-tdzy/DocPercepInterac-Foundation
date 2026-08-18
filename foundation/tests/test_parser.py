"""Unit tests for perception/parser.py — Geometry Layer (P1-03 acceptance criteria).

Docling removed 2026-08-11 — parser is now python-docx (DOCX) + pdfplumber
(PDF), both deterministic, no AI, no model loading.
"""
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from perception.parser import extract_geometry, parse_docx, parse_pdf

FIXTURES = Path(__file__).resolve().parent / "fixtures"


def test_parse_docx_returns_ordered_blocks_with_indices():
    blocks = parse_docx(str(FIXTURES / "fixture_bcdt.docx"))
    assert isinstance(blocks, list)
    assert len(blocks) > 0
    # Text non-emptiness and the paragraph-xor-cell invariant only hold for
    # the original two block kinds — the Comprehensive Document Perception
    # phase added kinds that are legitimate exceptions by design: an image
    # has no text at all, and headers/footers/footnotes/endnotes/comments
    # have neither a body paragraph_index nor a table_index (they live
    # outside body reading order entirely) — see perception/parser.py.
    text_blocks = [b for b in blocks if b["kind"] in ("paragraph", "table_cell")]
    assert text_blocks, "fixture should still contain at least one paragraph/table-cell block"
    assert all(b["text"].strip() for b in text_blocks)
    for b in text_blocks:
        is_para = b["paragraph_index"] is not None
        is_cell = b["table_index"] is not None
        assert is_para != is_cell


def test_parse_docx_keeps_empty_table_cells(tmp_path):
    """Empty table cells are real fill-in placeholders (e.g. a financial
    figure a mapping rule or a user is meant to write in), not noise —
    unlike empty paragraphs, they must still become a block so they can
    be shown/edited in the UI before anything has been written into them.
    """
    from docx import Document as DocxDocument

    doc = DocxDocument()
    table = doc.add_table(rows=1, cols=2)
    table.rows[0].cells[0].text = "Label"
    # cells[1] stays empty on purpose
    path = tmp_path / "doc.docx"
    doc.save(path)

    blocks = parse_docx(str(path))
    cell_blocks = [b for b in blocks if b["table_index"] is not None]
    assert len(cell_blocks) == 2

    empty_block = next(b for b in cell_blocks if b["col_index"] == 1)
    assert empty_block["text"] == ""
    assert empty_block["row_index"] == 0
    assert empty_block["table_index"] == 0


def test_parse_docx_under_60s_cpu():
    start = time.monotonic()
    parse_docx(str(FIXTURES / "fixture_bcdt.docx"))
    elapsed = time.monotonic() - start
    assert elapsed < 60, f"Parse took {elapsed:.1f}s, exceeds 60s CPU budget"


def test_parse_pdf_digital_returns_blocks_with_bbox():
    # fixture_report.pdf is scanned/image-only (no text layer) — use the
    # confirmed born-digital fixture instead (see STATUS.md).
    blocks = parse_pdf(str(FIXTURES / "fixture_report_2.pdf"))
    assert isinstance(blocks, list)
    assert len(blocks) > 0
    for b in blocks:
        assert b["page"] is not None and b["page"] >= 1
        assert b["bbox"] is not None and len(b["bbox"]) == 4


def test_parse_pdf_digital_has_realistic_multipage_diversity():
    # fixture_report_2.pdf is a 32-page real financial statement — this
    # asserts the diversity anchor_builder.py will need to handle: the same
    # header/footer text repeats verbatim across many pages (e.g. company
    # name, form number, boilerplate), so anchor resolution can't assume
    # text_fingerprint alone is unique — page context matters.
    blocks = parse_pdf(str(FIXTURES / "fixture_report_2.pdf"))
    pages = {b["page"] for b in blocks}
    assert len(pages) > 10, "expected a genuinely multi-page fixture"

    from collections import Counter

    text_counts = Counter(b["text"].strip() for b in blocks)
    repeated = {t: c for t, c in text_counts.items() if c >= 10}
    assert repeated, (
        "expected at least one line (e.g. a repeated page header/footer) "
        "to appear on 10+ pages — this is the duplicate-text case "
        "anchor_builder.py's Strategy 1 (style+fingerprint) must not treat "
        "as unique on its own"
    )


def test_parse_pdf_scanned_returns_no_crash_zero_blocks():
    # No text layer -> zero blocks is the correct deterministic result, not
    # a crash. OCR is a separate, not-yet-decided concern.
    blocks = parse_pdf(str(FIXTURES / "fixture_report.pdf"))
    assert isinstance(blocks, list)


def test_extract_geometry_dispatches_by_extension():
    assert extract_geometry(str(FIXTURES / "fixture_bcdt.docx")) == parse_docx(
        str(FIXTURES / "fixture_bcdt.docx")
    )


def test_extract_geometry_rejects_unsupported_extension(tmp_path):
    bogus = tmp_path / "file.txt"
    bogus.write_text("hello")
    try:
        extract_geometry(str(bogus))
        assert False, "expected ValueError"
    except ValueError:
        pass


def test_parse_xlsx_extracts_correct_geometry(tmp_path):
    import openpyxl
    file_path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Hello"
    ws["B2"] = "World"
    
    # Create named range
    wb.create_named_range("MyNamedRange", ws, "A1:A1")
    
    wb.save(file_path)
    wb.close()
    
    from perception.parser import parse_xlsx
    blocks = parse_xlsx(str(file_path))
    assert len(blocks) == 2
    assert blocks[0]["text"] == "Hello"
    assert blocks[0]["sheet_name"] == "TestSheet"
    assert blocks[0]["cell_address"] == "A1"
    assert blocks[0]["named_range"] == "MyNamedRange"
    
    assert blocks[1]["text"] == "World"
    assert blocks[1]["cell_address"] == "B2"
    assert blocks[1]["named_range"] is None
