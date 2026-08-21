"""
Detailed inspection of XLSX sheets, tables, and financial values
"""
from pathlib import Path
import sys
import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO_ROOT = Path(__file__).resolve().parents[3]
PATH_FARPT = REPO_ROOT / "anonymize client/Demo files/Demo files/FA&RPTS & Appendix I/FA&RPTs/HMV-FA&RPT FY2024.xlsx"
PATH_APP1 = REPO_ROOT / "anonymize client/Demo files/Demo files/FA&RPTS & Appendix I/Appendix I/HMV-25-Appendix I under D20 for FY2024-Final-W3103.xlsx"

wb_f = openpyxl.load_workbook(str(PATH_FARPT), data_only=True)
wb_a = openpyxl.load_workbook(str(PATH_APP1), data_only=True)

print("="*80)
print("FA&RPT FY2024 SHEETS & SAMPLES:")
for sname in wb_f.sheetnames:
    ws = wb_f[sname]
    print(f"\n--- Sheet: {sname} ({ws.max_row}x{ws.max_column}) ---")
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "") for c in range(1, min(8, ws.max_column + 1))]
        if any(row_vals):
            print(f"  Row {r:2d}: {' | '.join(row_vals)}")

print("\n"+"="*80)
print("APPENDIX I FY2024 SELECTED SHEETS & SAMPLES:")
for sname in ["I. Related parties", "III. Summary-RPTs", "IV. Segmented data", "Interest expenses", "Full Appendix I"]:
    if sname in wb_a.sheetnames:
        ws = wb_a[sname]
        print(f"\n--- Appendix I Sheet: {sname} ({ws.max_row}x{ws.max_column}) ---")
        for r in range(1, min(15, ws.max_row + 1)):
            row_vals = [str(ws.cell(r, c).value or "") for c in range(1, min(8, ws.max_column + 1))]
            if any(row_vals):
                print(f"  Row {r:2d}: {' | '.join(row_vals)}")
